"""Formula-bearing XLSX serialization for aperture U-value audit reports."""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from features.aperture_u_value.models import (
    ApertureReportElement,
    ApertureReportSection,
    ApertureUValueReport,
)
from features.aperture_u_value.report import (
    neutralize_spreadsheet_text,
    report_element_warning_text,
)
from features.aperture_u_value.units import (
    W_M2K_TO_BTU_HRFT2F,
    ExportUnitSystem,
    area_from_m2,
    area_label,
    frame_width_from_m,
    frame_width_label,
    heat_flow_label,
    length_from_m,
    length_label,
    psi_from_w_mk,
    psi_label,
    u_value_from_w_m2k,
    u_value_label,
)
from features.project_document.aperture_commands.models import APERTURE_SIDES

DETAIL_SHEET: Final = "Window Units"
SUMMARY_SHEET: Final = "Summary"
HEADER_ROW: Final = 9
FIRST_DATA_ROW: Final = HEADER_ROW + 1
UNFINISHED: Final = "UNFINISHED"

_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
_CALC_FILL = PatternFill("solid", fgColor="E9F4E2")
_UNFINISHED_FILL = PatternFill("solid", fgColor="FCE8E6")

_CALCULATED_KEYS = tuple(
    f"{side}_{suffix}"
    for side in APERTURE_SIDES
    for suffix in (
        "interior_length",
        "center_area",
        "corner_a",
        "corner_b",
        "frame_area",
        "q_frame",
        "q_spacer",
    )
) + (
    "interior_width",
    "interior_height",
    "glazing_area",
    "q_glazing",
    "q_frame_total",
    "q_spacer_total",
    "element_u",
)


@dataclass(frozen=True)
class _Layout:
    headers: tuple[str, ...]
    columns: dict[str, int]

    def col(self, key: str) -> int:
        return self.columns[key]

    def ref(self, key: str, row: int, *, absolute: bool = False) -> str:
        letter = get_column_letter(self.col(key))
        return f"${letter}${row}" if absolute else f"{letter}{row}"


def render_aperture_u_value_xlsx(
    report: ApertureUValueReport,
    *,
    units: ExportUnitSystem,
    generated_at: datetime,
) -> bytes:
    """Build a workbook whose formulas reproduce PHN's rounded SI results."""
    workbook = Workbook()
    detail = workbook.active
    detail.title = DETAIL_SHEET
    summary = workbook.create_sheet(SUMMARY_SHEET)
    layout = _layout(units)
    last_data_row = HEADER_ROW + sum(len(section.elements) for section in report.apertures)

    _write_provenance(detail, report, generated_at)
    _write_headers(detail, layout)
    first_row_by_aperture: dict[str, int] = {}
    row_number = FIRST_DATA_ROW
    for section in report.apertures:
        aperture_last_data_row = row_number + len(section.elements) - 1
        for element in section.elements:
            rollup_row = first_row_by_aperture.setdefault(
                section.aperture_type_id,
                row_number,
            )
            _write_detail_row(
                detail,
                layout,
                row_number,
                section,
                element,
                units,
                rollup_row,
                aperture_last_data_row,
            )
            row_number += 1

    _finish_detail_sheet(detail, layout, last_data_row)
    _write_summary(
        summary,
        report,
        layout,
        units,
        first_row_by_aperture,
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _layout(units: ExportUnitSystem) -> _Layout:
    length_unit = length_label(units)
    width_unit = frame_width_label(units)
    area_unit = area_label(units)
    u_unit = u_value_label(units)
    psi_unit = psi_label(units)
    heat_unit = heat_flow_label(units)
    definitions: list[tuple[str, str]] = [
        ("aperture_id", "Aperture ID"),
        ("aperture_name", "Aperture Name"),
        ("element_id", "Element ID"),
        ("element_name", "Element Name"),
        ("grid", "Grid Position"),
        ("status", "Status"),
        ("width", f"Width [{length_unit}]"),
        ("height", f"Height [{length_unit}]"),
        ("area", f"Element Area [{area_unit}]"),
        ("glazing_id", "Glazing ID"),
        ("glazing_name", "Glazing Name"),
        ("glazing_u", f"Glazing U [{u_unit}]"),
        ("shgc", "Glazing SHGC"),
    ]
    for side in APERTURE_SIDES:
        definitions.extend(
            [
                (f"{side}_frame_name", f"{side.title()} Frame"),
                (
                    f"{side}_frame_width",
                    f"{side.title()} Frame Width [{width_unit}]",
                ),
                (f"{side}_frame_u", f"{side.title()} Frame U [{u_unit}]"),
                (f"{side}_psi_g", f"{side.title()} Ψ-g [{psi_unit}]"),
                (
                    f"{side}_psi_install",
                    f"{side.title()} Ψ-install [{psi_unit}] (excluded from U-w)",
                ),
                (
                    f"{side}_edge_length",
                    f"{side.title()} Edge Length [{length_unit}]",
                ),
                (
                    f"{side}_interior_length",
                    f"{side.title()} Interior Length [{length_unit}]",
                ),
                (
                    f"{side}_center_area",
                    f"{side.title()} Center-Strip Area [{area_unit}]",
                ),
                (
                    f"{side}_corner_a",
                    f"{side.title()} Half-Corner A [{area_unit}]",
                ),
                (
                    f"{side}_corner_b",
                    f"{side.title()} Half-Corner B [{area_unit}]",
                ),
                (
                    f"{side}_frame_area",
                    f"{side.title()} Frame Area [{area_unit}]",
                ),
                (
                    f"{side}_q_frame",
                    f"{side.title()} Q-frame [{heat_unit}]",
                ),
                (
                    f"{side}_q_spacer",
                    f"{side.title()} Q-spacer [{heat_unit}]",
                ),
            ]
        )
    definitions.extend(
        [
            ("interior_width", f"Interior Width [{length_unit}]"),
            ("interior_height", f"Interior Height [{length_unit}]"),
            ("glazing_area", f"Glazing Area [{area_unit}]"),
            ("q_glazing", f"Q-glazing [{heat_unit}]"),
            ("q_frame_total", f"Q-frame Total [{heat_unit}]"),
            ("q_spacer_total", f"Q-spacer Total [{heat_unit}]"),
            ("element_u", f"Element U [{u_unit}]"),
            ("aperture_area", f"Aperture Area [{area_unit}]"),
            ("aperture_u", f"Aperture U-w [{u_unit}]"),
            ("aperture_shgc", "Aperture SHGC [glazing-area weighted]"),
            ("warnings", "Warnings"),
        ]
    )
    return _Layout(
        headers=tuple(label for _, label in definitions),
        columns={key: index for index, (key, _) in enumerate(definitions, start=1)},
    )


def _write_provenance(
    sheet: Worksheet,
    report: ApertureUValueReport,
    generated_at: datetime,
) -> None:
    sheet["A1"] = "PH-Navigator Aperture U-Value Detail Report"
    sheet["A1"].font = Font(bold=True, size=14)
    rows = (
        ("Project", report.provenance.project_name),
        ("BT Number", report.provenance.bt_number),
        ("Version", report.provenance.version_label),
        ("Generated", generated_at.isoformat()),
        (
            "Convention",
            (f"{report.provenance.generated_note} · shared mullions are charged fully to each adjacent element"),
        ),
        (
            "Rollup",
            (
                "Aperture U-w includes unfinished elements as U = 0."
                if any(section.unfinished_count for section in report.apertures)
                else "All exported glazed elements are complete."
            ),
        ),
        (
            "SHGC",
            "Aperture SHGC is weighted by positive glazing area; missing g-values are excluded.",
        ),
    )
    for row_number, (label, value) in enumerate(rows, start=2):
        sheet.cell(row=row_number, column=1, value=label).font = Font(bold=True)
        sheet.cell(
            row=row_number,
            column=2,
            value=neutralize_spreadsheet_text(value),
        )


def _write_headers(sheet: Worksheet, layout: _Layout) -> None:
    for column, header in enumerate(layout.headers, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_detail_row(
    sheet: Worksheet,
    layout: _Layout,
    row: int,
    section: ApertureReportSection,
    element: ApertureReportElement,
    units: ExportUnitSystem,
    rollup_row: int,
    aperture_last_data_row: int,
) -> None:
    edge_by_side = {edge.side: edge for edge in element.edges}
    _set(sheet, layout, row, "aperture_id", section.aperture_type_id)
    _set_text(sheet, layout, row, "aperture_name", section.name)
    _set(sheet, layout, row, "element_id", element.element_id)
    _set_text(sheet, layout, row, "element_name", element.element_name)
    _set(sheet, layout, row, "grid", element.grid_label)
    _set(sheet, layout, row, "status", UNFINISHED if element.unfinished else "COMPLETE")
    _set(sheet, layout, row, "width", length_from_m(element.width_m, units))
    _set(sheet, layout, row, "height", length_from_m(element.height_m, units))
    _set(sheet, layout, row, "area", area_from_m2(element.area_m2, units))
    _set(sheet, layout, row, "glazing_id", element.glazing_id)
    _set_text(sheet, layout, row, "glazing_name", element.glazing_name)
    _set(
        sheet,
        layout,
        row,
        "glazing_u",
        u_value_from_w_m2k(element.glazing_u_w_m2k, units),
    )
    _set(
        sheet,
        layout,
        row,
        "shgc",
        element.glazing_g_value,
    )

    for side in APERTURE_SIDES:
        edge = edge_by_side[side]
        _set_text(sheet, layout, row, f"{side}_frame_name", edge.frame_name)
        _set(
            sheet,
            layout,
            row,
            f"{side}_frame_width",
            frame_width_from_m(edge.width_m, units),
        )
        _set(
            sheet,
            layout,
            row,
            f"{side}_frame_u",
            u_value_from_w_m2k(edge.u_value_w_m2k, units),
        )
        _set(
            sheet,
            layout,
            row,
            f"{side}_psi_g",
            psi_from_w_mk(edge.psi_g_w_mk, units),
        )
        _set(
            sheet,
            layout,
            row,
            f"{side}_psi_install",
            psi_from_w_mk(edge.psi_install_w_mk, units),
        )
        _set(
            sheet,
            layout,
            row,
            f"{side}_edge_length",
            length_from_m(edge.edge_length_m, units),
        )

    if element.unfinished:
        for key in _CALCULATED_KEYS:
            _set(sheet, layout, row, key, UNFINISHED)
    else:
        _write_calculation_formulas(sheet, layout, row, units)

    if row == rollup_row:
        _set(
            sheet,
            layout,
            row,
            "aperture_area",
            _aperture_area_formula(
                layout,
                row,
                aperture_last_data_row,
            ),
        )
        _set(
            sheet,
            layout,
            row,
            "aperture_u",
            _aperture_u_formula(
                layout,
                row,
                aperture_last_data_row,
                units,
            ),
        )
        _set(
            sheet,
            layout,
            row,
            "aperture_shgc",
            _aperture_shgc_formula(
                layout,
                row,
                aperture_last_data_row,
            ),
        )
    else:
        for key in ("aperture_area", "aperture_u", "aperture_shgc"):
            _set(sheet, layout, row, key, f"={layout.ref(key, rollup_row)}")
    _set_text(
        sheet,
        layout,
        row,
        "warnings",
        report_element_warning_text(element, section),
    )

    if element.unfinished:
        for cell in sheet[row]:
            cell.fill = _UNFINISHED_FILL
    else:
        for key in _CALCULATED_KEYS:
            sheet.cell(row=row, column=layout.col(key)).fill = _CALC_FILL


def _write_calculation_formulas(
    sheet: Worksheet,
    layout: _Layout,
    row: int,
    units: ExportUnitSystem,
) -> None:
    widths_per_length = 1000 if units == "SI" else 12
    adjacent = {
        "top": ("left", "right", "width"),
        "right": ("top", "bottom", "height"),
        "bottom": ("left", "right", "width"),
        "left": ("top", "bottom", "height"),
    }
    for side, (adjacent_a, adjacent_b, dimension) in adjacent.items():
        width = layout.ref(f"{side}_frame_width", row)
        width_in_length = f"({width}/{widths_per_length})"
        adjacent_a_width = layout.ref(f"{adjacent_a}_frame_width", row)
        adjacent_b_width = layout.ref(f"{adjacent_b}_frame_width", row)
        interior = (
            f"={layout.ref(dimension, row)}"
            f"-{adjacent_a_width}/{widths_per_length}"
            f"-{adjacent_b_width}/{widths_per_length}"
        )
        _set(sheet, layout, row, f"{side}_interior_length", interior)
        interior_ref = layout.ref(f"{side}_interior_length", row)
        _set(
            sheet,
            layout,
            row,
            f"{side}_center_area",
            f"={width_in_length}*{interior_ref}",
        )
        _set(
            sheet,
            layout,
            row,
            f"{side}_corner_a",
            (f"={width_in_length}*({adjacent_a_width}/{widths_per_length})/2"),
        )
        _set(
            sheet,
            layout,
            row,
            f"{side}_corner_b",
            (f"={width_in_length}*({adjacent_b_width}/{widths_per_length})/2"),
        )
        _set(
            sheet,
            layout,
            row,
            f"{side}_frame_area",
            (
                f"=SUM({layout.ref(f'{side}_center_area', row)},"
                f"{layout.ref(f'{side}_corner_a', row)},"
                f"{layout.ref(f'{side}_corner_b', row)})"
            ),
        )
        _set(
            sheet,
            layout,
            row,
            f"{side}_q_frame",
            (f"={layout.ref(f'{side}_frame_area', row)}*{layout.ref(f'{side}_frame_u', row)}"),
        )
        _set(
            sheet,
            layout,
            row,
            f"{side}_q_spacer",
            (f"={layout.ref(f'{side}_interior_length', row)}*{layout.ref(f'{side}_psi_g', row)}"),
        )

    _set(
        sheet,
        layout,
        row,
        "interior_width",
        (
            f"={layout.ref('width', row)}"
            f"-{layout.ref('left_frame_width', row)}/{widths_per_length}"
            f"-{layout.ref('right_frame_width', row)}/{widths_per_length}"
        ),
    )
    _set(
        sheet,
        layout,
        row,
        "interior_height",
        (
            f"={layout.ref('height', row)}"
            f"-{layout.ref('top_frame_width', row)}/{widths_per_length}"
            f"-{layout.ref('bottom_frame_width', row)}/{widths_per_length}"
        ),
    )
    _set(
        sheet,
        layout,
        row,
        "glazing_area",
        (f"={layout.ref('interior_width', row)}*{layout.ref('interior_height', row)}"),
    )
    _set(
        sheet,
        layout,
        row,
        "q_glazing",
        f"={layout.ref('glazing_area', row)}*{layout.ref('glazing_u', row)}",
    )
    q_frame_refs = ",".join(layout.ref(f"{side}_q_frame", row) for side in APERTURE_SIDES)
    q_spacer_refs = ",".join(layout.ref(f"{side}_q_spacer", row) for side in APERTURE_SIDES)
    _set(
        sheet,
        layout,
        row,
        "q_frame_total",
        f"=SUM({q_frame_refs})",
    )
    _set(
        sheet,
        layout,
        row,
        "q_spacer_total",
        f"=SUM({q_spacer_refs})",
    )
    raw_u = (
        f"({layout.ref('q_glazing', row)}"
        f"+{layout.ref('q_frame_total', row)}"
        f"+{layout.ref('q_spacer_total', row)})"
        f"/({layout.ref('width', row)}*{layout.ref('height', row)})"
    )
    element_formula = (
        f"=ROUND({raw_u},4)" if units == "SI" else (f"=ROUND(({raw_u})/{W_M2K_TO_BTU_HRFT2F},4)*{W_M2K_TO_BTU_HRFT2F}")
    )
    _set(sheet, layout, row, "element_u", element_formula)


def _aperture_area_formula(
    layout: _Layout,
    first_data_row: int,
    last_data_row: int,
) -> str:
    areas = _range(layout, "area", first_data_row, last_data_row)
    return f"=SUM({areas})"


def _aperture_u_formula(
    layout: _Layout,
    first_data_row: int,
    last_data_row: int,
    units: ExportUnitSystem,
) -> str:
    areas = _range(layout, "area", first_data_row, last_data_row)
    element_u = _range(
        layout,
        "element_u",
        first_data_row,
        last_data_row,
    )
    weighted = f"SUMPRODUCT({areas},{element_u})/SUM({areas})"
    if units == "SI":
        return f"=ROUND({weighted},4)"
    return f"=ROUND(({weighted})/{W_M2K_TO_BTU_HRFT2F},4)*{W_M2K_TO_BTU_HRFT2F}"


def _aperture_shgc_formula(
    layout: _Layout,
    first_data_row: int,
    last_data_row: int,
) -> str:
    glazing_areas = _range(
        layout,
        "glazing_area",
        first_data_row,
        last_data_row,
    )
    shgc = _range(layout, "shgc", first_data_row, last_data_row)
    numerator = f"SUMPRODUCT({glazing_areas},{shgc})"
    denominator = f"SUMPRODUCT({glazing_areas},--ISNUMBER({shgc}))"
    return f'=IF({denominator}=0,"",{numerator}/{denominator})'


def _range(
    layout: _Layout,
    key: str,
    first_data_row: int,
    last_data_row: int,
) -> str:
    letter = get_column_letter(layout.col(key))
    return f"${letter}${first_data_row}:${letter}${last_data_row}"


def _finish_detail_sheet(
    sheet: Worksheet,
    layout: _Layout,
    last_data_row: int,
) -> None:
    sheet.freeze_panes = f"A{FIRST_DATA_ROW}"
    if last_data_row >= FIRST_DATA_ROW:
        sheet.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(layout.headers))}{last_data_row}"
        for row in sheet.iter_rows(
            min_row=FIRST_DATA_ROW,
            max_row=last_data_row,
            min_col=layout.col("width"),
            max_col=layout.col("aperture_shgc"),
        ):
            for cell in row:
                if cell.value is not None and cell.value != UNFINISHED:
                    cell.number_format = "0.###############"
    for index, header in enumerate(layout.headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(len(header) + 2, 12),
            28,
        )
    sheet.column_dimensions[get_column_letter(layout.col("warnings"))].width = 56


def _write_summary(
    sheet: Worksheet,
    report: ApertureUValueReport,
    layout: _Layout,
    units: ExportUnitSystem,
    first_row_by_aperture: dict[str, int],
) -> None:
    headers = (
        "Aperture ID",
        "Aperture Name",
        f"Area [{area_label(units)}]",
        f"U-w [{u_value_label(units)}]",
        "SHGC [glazing-area weighted]",
        "Unfinished Elements",
        "Void Panels Excluded",
    )
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    for row, section in enumerate(report.apertures, start=2):
        detail_row = first_row_by_aperture.get(section.aperture_type_id)
        sheet.cell(row=row, column=1, value=section.aperture_type_id)
        sheet.cell(
            row=row,
            column=2,
            value=neutralize_spreadsheet_text(section.name),
        )
        if detail_row is None:
            sheet.cell(row=row, column=3, value="=0")
            sheet.cell(row=row, column=4, value="=0")
            sheet.cell(row=row, column=5, value='=""')
        else:
            quoted_sheet = f"'{DETAIL_SHEET}'"
            sheet.cell(
                row=row,
                column=3,
                value=f"={quoted_sheet}!{layout.ref('aperture_area', detail_row)}",
            )
            sheet.cell(
                row=row,
                column=4,
                value=f"={quoted_sheet}!{layout.ref('aperture_u', detail_row)}",
            )
            sheet.cell(
                row=row,
                column=5,
                value=f"={quoted_sheet}!{layout.ref('aperture_shgc', detail_row)}",
            )
        sheet.cell(row=row, column=6, value=section.unfinished_count)
        sheet.cell(row=row, column=7, value=section.void_count)
        for column in (3, 4, 5):
            sheet.cell(row=row, column=column).number_format = "0.###############"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{max(1, len(report.apertures) + 1)}"
    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 24


def _set(
    sheet: Worksheet,
    layout: _Layout,
    row: int,
    key: str,
    value: object,
) -> None:
    sheet.cell(row=row, column=layout.col(key), value=value)


def _set_text(
    sheet: Worksheet,
    layout: _Layout,
    row: int,
    key: str,
    value: str | None,
) -> None:
    sheet.cell(
        row=row,
        column=layout.col(key),
        value=neutralize_spreadsheet_text(value) if value is not None else None,
    )

"""OneBuilding EPW catalog fetch, nearest lookup, and zip extraction."""

from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from dataclasses import replace as dataclass_replace
from datetime import UTC, datetime, timedelta
from functools import lru_cache
from urllib.parse import urljoin

import certifi
import httpx
from openpyxl import load_workbook

from config import settings
from features.climate.proximity import haversine_miles

ONEBUILDING_BASE_URL = "https://climate.onebuilding.org/"
DEFAULT_CATALOG_URLS = (
    "https://climate.onebuilding.org/sources/Region1_Africa_TMYx_EPW_Processing_locations.xlsx",
    "https://climate.onebuilding.org/sources/Region2_Asia_TMYx_EPW_Processing_locations.xlsx",
    "https://climate.onebuilding.org/sources/Region3_South_America_TMYx_EPW_Processing_locations.xlsx",
    "https://climate.onebuilding.org/sources/Region4_USA_TMYx_EPW_Processing_locations.xlsx",
    "https://climate.onebuilding.org/sources/Region4_Canada_TMYx_EPW_Processing_locations.xlsx",
    "https://climate.onebuilding.org/sources/Region4_NA_CA_Caribbean_TMYx_EPW_Processing_locations.xlsx",
    "https://climate.onebuilding.org/sources/Region5_Southwest_Pacific_TMYx_EPW_Processing_locations.xlsx",
    "https://climate.onebuilding.org/sources/Region6_Europe_TMYx_EPW_Processing_locations.xlsx",
    "https://climate.onebuilding.org/sources/Region7_Antarctica_TMYx_EPW_Processing_locations.xlsx",
)
_CATALOG_TTL = timedelta(hours=24)
_YEAR_SPAN_RE = re.compile(r"\.(\d{4})-(\d{4})\.zip$", re.IGNORECASE)
# OneBuilding encodes the dataset version in the zip name after the WMO, e.g.
# ``…744104_TMYx.2009-2023.zip`` (type + period) or ``…744104_TMY3.zip`` (type).
_VERSION_RE = re.compile(r"_(?P<type>[A-Za-z0-9]+)(?:\.(?P<period>\d{4}-\d{4}))?\.zip$", re.IGNORECASE)


@dataclass(frozen=True)
class EpwCatalogEntry:
    country: str | None
    region: str | None
    name: str
    wmo: str | None
    source_data: str | None
    latitude: float
    longitude: float
    elevation_m: float | None
    time_zone_offset_hours: float | None
    url: str
    distance_mi: float | None = None


@dataclass(frozen=True)
class EpwZipPayload:
    entry: EpwCatalogEntry
    epw_name: str
    epw_bytes: bytes
    stat_name: str | None
    stat_text: str | None


def nearest_epw_entry(latitude: float, longitude: float) -> EpwCatalogEntry | None:
    entries = load_epw_catalog(catalog_urls())
    if not entries:
        return None
    nearest = min(
        entries,
        key=lambda entry: (
            round(haversine_miles(latitude, longitude, entry.latitude, entry.longitude), 6),
            _recency_rank(entry),
        ),
    )
    return dataclass_replace(
        nearest,
        distance_mi=haversine_miles(latitude, longitude, nearest.latitude, nearest.longitude),
    )


def nearest_epw_entries(
    latitude: float, longitude: float, *, country: str | None = None, limit: int | None
) -> list[EpwCatalogEntry]:
    """Catalog entries nearest a point, nearest-first, each stamped with distance.

    Generalizes :func:`nearest_epw_entry` to the top ``limit`` results (``None``
    = no cap); an optional ``country`` narrows the search (the any-state weather
    picker mode).
    """
    entries = [entry for entry in load_epw_catalog(catalog_urls()) if _matches(entry.country, country)]
    return _rank_nearest(entries, latitude, longitude, limit)


def epw_entries_for_region(
    *, country: str | None, region: str | None, latitude: float, longitude: float, limit: int | None
) -> list[EpwCatalogEntry]:
    """Catalog entries for a country/region (state), nearest-first with distance.

    ``limit`` of ``None`` returns the whole state (the picker shows every dataset
    version, so a state's full roster is the intent)."""
    entries = [
        entry
        for entry in load_epw_catalog(catalog_urls())
        if _matches(entry.country, country) and _matches(entry.region, region)
    ]
    return _rank_nearest(entries, latitude, longitude, limit)


def find_entry_by_url(url: str) -> EpwCatalogEntry | None:
    """Resolve a catalog entry from its zip URL (for a from-catalog attach)."""
    return next((entry for entry in load_epw_catalog(catalog_urls()) if entry.url == url), None)


def epw_version_label(url: str) -> str:
    """Human label for an EPW file's dataset version, parsed from its filename.

    OneBuilding encodes the methodology and period in the zip name, e.g.
    ``…_TMYx.2009-2023.zip`` → ``TMYx 2009–2023`` and ``…_TMY3.zip`` → ``TMY3``.
    A station has several such files; the picker lists them all, labelled, so the
    user can tell the versions apart. Falls back to ``EPW`` off-convention."""
    match = _VERSION_RE.search(url)
    if match is None:
        return "EPW"
    period = match.group("period")
    dataset_type = match.group("type")
    return f"{dataset_type} {period.replace('-', '–')}" if period else dataset_type


def _rank_nearest(
    entries: list[EpwCatalogEntry], latitude: float, longitude: float, limit: int | None
) -> list[EpwCatalogEntry]:
    """Order catalog entries for the picker and stamp each with its great-circle
    ``distance_mi``.

    First reduce to one row per (station, dataset type) — keeping the most recent
    period when a type is dated (``TMYx 2007-2021``/``2009-2023``/``2011-2025`` →
    just ``2011-2025``) while preserving distinct methodologies (TMYx, TMY3, …).
    Then order stations nearest-first, keeping a station's remaining versions
    adjacent and most-recent-first rather than interleaved with a neighbour in
    the same distance band. ``limit`` caps the row count; ``None`` means no cap
    (region mode shows a whole state)."""
    reduced = _collapse_to_recent_per_type(entries)
    decorated = [
        (haversine_miles(latitude, longitude, entry.latitude, entry.longitude), _station_key(entry), entry)
        for entry in reduced
    ]
    nearest_per_station: dict[tuple[str, str, str], float] = {}
    for distance, key, _entry in decorated:
        nearest_per_station[key] = min(distance, nearest_per_station.get(key, distance))
    decorated.sort(key=lambda row: (round(nearest_per_station[row[1]], 6), row[1], _recency_rank(row[2])))
    capped = decorated if limit is None else decorated[: max(limit, 0)]
    return [dataclass_replace(entry, distance_mi=distance) for distance, _key, entry in capped]


def _collapse_to_recent_per_type(entries: list[EpwCatalogEntry]) -> list[EpwCatalogEntry]:
    """One entry per (station, dataset type): when a type carries dated periods,
    keep only the most recent. OneBuilding lists a station's TMYx once per rolling
    period (``…2007-2021``, ``…2009-2023``, ``…2011-2025``) — only the newest is
    useful — but a different methodology (TMY3, …) is a distinct choice and is
    kept. Selection is by :func:`_recency_rank` (dated-most-recent wins)."""
    best: dict[tuple[tuple[str, str, str], str], EpwCatalogEntry] = {}
    for entry in entries:
        key = (_station_key(entry), _dataset_type(entry.url))
        incumbent = best.get(key)
        if incumbent is None or _recency_rank(entry) < _recency_rank(incumbent):
            best[key] = entry
    return list(best.values())


def _station_key(entry: EpwCatalogEntry) -> tuple[str, str, str]:
    """Identity of the physical station an entry belongs to — its catalog
    country/region/name. Stable across a station's version rows, whose
    coordinates can drift a few hundred metres between source periods."""
    return (
        (entry.country or "").strip().casefold(),
        (entry.region or "").strip().casefold(),
        entry.name.strip().casefold(),
    )


def _dataset_type(url: str) -> str:
    """The methodology token of an EPW file (``TMYx``, ``TMY3``, …), casefolded
    for grouping; ``""`` when the filename is off-convention."""
    match = _VERSION_RE.search(url)
    return match.group("type").casefold() if match else ""


def _matches(value: str | None, target: str | None) -> bool:
    """Case-insensitive equality; a ``None`` target matches anything (no filter)."""
    if target is None:
        return True
    return value is not None and value.strip().casefold() == target.strip().casefold()


def download_epw_zip(entry: EpwCatalogEntry) -> EpwZipPayload:
    with httpx.Client(
        timeout=settings.location_derive_timeout_seconds,
        verify=certifi.where(),
        follow_redirects=True,
    ) as client:
        response = client.get(entry.url)
        response.raise_for_status()
    return epw_zip_payload(entry, response.content)


def epw_zip_payload(entry: EpwCatalogEntry, raw_zip: bytes) -> EpwZipPayload:
    with zipfile.ZipFile(io.BytesIO(raw_zip)) as archive:
        epw_name = next((name for name in archive.namelist() if name.lower().endswith(".epw")), None)
        if epw_name is None:
            raise ValueError("epw_zip_missing_epw")
        stat_name = next((name for name in archive.namelist() if name.lower().endswith(".stat")), None)
        stat_text = archive.read(stat_name).decode("utf-8-sig", errors="replace") if stat_name else None
        return EpwZipPayload(
            entry=entry,
            epw_name=epw_name.rsplit("/", maxsplit=1)[-1],
            epw_bytes=archive.read(epw_name),
            stat_name=stat_name.rsplit("/", maxsplit=1)[-1] if stat_name else None,
            stat_text=stat_text,
        )


def load_epw_catalog(urls: tuple[str, ...]) -> tuple[EpwCatalogEntry, ...]:
    return _cached_catalog(urls, _ttl_bucket())


def catalog_urls() -> tuple[str, ...]:
    configured = [item.strip() for item in settings.epw_catalog_urls.split(",") if item.strip()]
    return tuple(configured) if configured else DEFAULT_CATALOG_URLS


@lru_cache(maxsize=4)
def _cached_catalog(urls: tuple[str, ...], _ttl: int) -> tuple[EpwCatalogEntry, ...]:
    entries: list[EpwCatalogEntry] = []
    with httpx.Client(
        timeout=settings.location_derive_timeout_seconds,
        verify=certifi.where(),
        follow_redirects=True,
    ) as client:
        for url in urls:
            try:
                response = client.get(url)
                response.raise_for_status()
            except httpx.HTTPError:
                continue
            entries.extend(parse_epw_catalog_xlsx(response.content, base_url=url))
    return tuple(entries)


def parse_epw_catalog_xlsx(raw: bytes, *, base_url: str = ONEBUILDING_BASE_URL) -> list[EpwCatalogEntry]:
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip().lower() if value is not None else "" for value in next(rows)]
    index = {header: idx for idx, header in enumerate(headers)}
    entries: list[EpwCatalogEntry] = []
    for row in rows:
        latitude = _float_at(row, index, "latitude (n+/s-)")
        longitude = _float_at(row, index, "longitude (e+/w-)")
        url = _text_at(row, index, "url")
        name = _text_at(row, index, "city/station")
        if latitude is None or longitude is None or not url or not name:
            continue
        entries.append(
            EpwCatalogEntry(
                country=_text_at(row, index, "country"),
                region=_text_at(row, index, "state"),
                name=name,
                wmo=_text_at(row, index, "wmo"),
                source_data=_text_at(row, index, "source data"),
                latitude=latitude,
                longitude=longitude,
                elevation_m=_float_at(row, index, "elevation (m)"),
                time_zone_offset_hours=_float_at(row, index, "time zone (gmt +/-)"),
                url=urljoin(base_url, url),
            )
        )
    return entries


def _recency_rank(entry: EpwCatalogEntry) -> tuple[int, int, int, str]:
    """Most-recent-first ordering key for a station's dataset versions: a dated
    file beats an undated one, then the latest period end-year wins, then
    SRC-backed source quality, then URL for determinism. Drives both the file the
    picker keeps per (station, type) and the distance tie-break when auto-picking
    the nearest station."""
    match = _YEAR_SPAN_RE.search(entry.url)
    end_year = int(match.group(2)) if match else 0
    source_penalty = 0 if (entry.source_data or "").startswith("SRC") else 1
    return (0 if match else 1, -end_year, source_penalty, entry.url)


def _text_at(row: tuple[object, ...], index: dict[str, int], header: str) -> str | None:
    position = index.get(header)
    if position is None or position >= len(row):
        return None
    value = row[position]
    text = str(value).strip() if value is not None else ""
    return text or None


def _float_at(row: tuple[object, ...], index: dict[str, int], header: str) -> float | None:
    text = _text_at(row, index, header)
    if text is None:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _ttl_bucket() -> int:
    return int(datetime.now(tz=UTC).timestamp() // _CATALOG_TTL.total_seconds())

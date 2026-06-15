"""Importer for the PHI/PHPP 10.6 climate-library workbook.

PHI ships its climate library *embedded* in a live PHPP ``Climate`` worksheet
(``.xlsx``), not as a clean per-location table. This module is the PHI *process*
parser: it walks the embedded library and turns each dataset row into a
:class:`ClimateRecord`. Like :mod:`features.climate.importers.phius`, it is
pure — no database, no object store; the bundle writer
(:mod:`features.climate.processing`) and the seed step
(:mod:`features.climate.seeding`) own everything downstream.

Recovering the column semantics is the load-bearing risk: the library columns
are unlabeled formula outputs, and a mis-mapped column silently seeds wrong
climate for ~1000 locations. The map below was reverse-engineered from the
10.6 sample workbook and **validated against ground truth** — decoded
coordinates match the real cities (e.g. Birmingham AL → 33.5°N/-86.92°,
Rochester NY → 43.12°N/-77.68°) and every row satisfies the physical
invariants dewpoint ≤ air temp, sky ≤ air temp, and global ≥ directional
radiation in summer.

The recovered layout of the ``Climate`` worksheet (1-based columns):

- A header band (rows 1-242) holds the single *active* climate display and the
  cascading-dropdown helper lists; it is skipped. The embedded library begins
  at row 243.
- Each **library row** is a dataset iff column ``K`` ("anzeigen?", show-flag)
  is ``"x"``. That flag excludes the ``aktuell`` active-selection echo row
  (``K = "-"``) and the import placeholders (``K`` empty); empty user-defined
  ``ud---NN`` template slots carry the flag but no data and are skipped on the
  absence of coordinates / an air-temperature series.
- Identity + scalars: ``I`` PHPP picker key (→ ``station_id``), ``N`` country,
  ``O`` region, ``P`` location name, ``S`` latitude, ``T`` longitude,
  ``U`` elevation (m, may be blank), ``V`` summer temperature swing (K).
- Eight consecutive **12-month series** start at column ``W`` (Jan…Dec each),
  in this fixed order: air temperature, North / East / South / West / global
  horizontal radiation, dewpoint, sky temperature. Dewpoint and sky are absent
  for some locations and default to a zero vector (as in the Phius importer).
- Four **peak/design conditions** follow: two heating (``DO…DT``, ``DU…DZ``)
  then two cooling (``EA…EG``, ``EH…EN``); each carries air temp + five
  radiation columns, and the cooling pair carries a trailing dewpoint column.
  A condition the source leaves blank (commonly the second cooling load)
  defaults to zeros.
- Trailing columns (MIN/MAX temp, wind, PER factors) are PHPP bookkeeping, not
  ``ClimateRecord`` fields, and are ignored.

Source quirks are imported faithfully, not "corrected": a handful of datasets
repeat December from November in the dewpoint/sky series, which can put
dewpoint marginally above air temp for that month — that is the licensed
source's value, and we serialize it verbatim.
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

import openpyxl

from features.climate.record import (
    ClimateAux,
    ClimateData,
    ClimateLocation,
    ClimateMonthlyRadiation,
    ClimateMonthlyTemps,
    ClimatePeakLoad,
    ClimatePeakLoads,
    ClimatePhppCodes,
    ClimateRecord,
)

_SHEET = "Climate"
_MONTHS = 12
_ZEROS = [0.0] * _MONTHS

# Library identity / scalar columns (1-based, as in the worksheet).
_COL_SHOW = 11  # K  "anzeigen?" — "x" marks a real, selectable dataset
_COL_KEY = 9  # I  PHPP picker key, e.g. "US0056b" → station_id
_COL_COUNTRY = 14  # N  "Land"
_COL_REGION = 15  # O  "Region"
_COL_NAME = 16  # P  "Standort" (location name)
_COL_LATITUDE = 19  # S  "Breite °"
_COL_LONGITUDE = 20  # T  "Länge °"
_COL_ELEVATION = 21  # U  "Meereshöhe m"
_COL_SUMMER_SWING = 22  # V  "Temperaturschwankung Sommer"

# The eight 12-month series run consecutively from column W (23). The order is
# fixed by the worksheet and verified against ground truth (see module docs).
_SERIES_START = 23
_SERIES_ORDER = ("air_c", "north", "east", "south", "west", "glob", "dewpoint_c", "sky_c")

# Four peak/design conditions follow the monthly blocks: (record field, first
# column, carries a trailing dewpoint column). Heating pair first, then cooling.
_PEAK_BLOCKS: tuple[tuple[str, int, bool], ...] = (
    ("heat_load_1", 119, False),  # DO..DT
    ("heat_load_2", 125, False),  # DU..DZ
    ("cooling_load_1", 131, True),  # EA..EF (+EG dewpoint)
    ("cooling_load_2", 138, True),  # EH..EM (+EN dewpoint)
)


class PhiParseError(ValueError):
    """Raised when the PHPP workbook does not match the expected layout."""


def iter_phi_records(root: Path) -> Iterator[ClimateRecord]:
    """Yield a record for every dataset in every PHPP workbook under ``root``.

    The registry hands the *process* step a source *tree*; PHI ships a single
    ``.xlsx``, so we walk the tree for workbooks (sorted, stable) and parse the
    embedded library out of each.
    """
    for path in sorted(root.rglob("*.xlsx")):
        yield from parse_phi_workbook(path)


def parse_phi_workbook(path: Path) -> Iterator[ClimateRecord]:
    """Yield a record for every selectable dataset embedded in one workbook.

    Non-dataset rows (the header band, the active-selection echo, and empty
    ``ud---NN`` template slots) are skipped; a row whose layout is genuinely
    malformed raises :class:`PhiParseError`.
    """
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if _SHEET not in workbook.sheetnames:
            raise PhiParseError(f"{path.name}: no {_SHEET!r} worksheet (found {workbook.sheetnames}).")
        # Read the sheet eagerly so the file handle is released before we start
        # yielding — a caller that abandons the generator must not leak the lock.
        rows = list(workbook[_SHEET].iter_rows(values_only=True))
    finally:
        workbook.close()

    for row in rows:
        record = _record_from_row(row)
        if record is not None:
            yield record


def _record_from_row(row: tuple[object, ...]) -> ClimateRecord | None:
    """Build one record from a worksheet row, or ``None`` if it is not a dataset."""
    if _cell(row, _COL_SHOW) != "x":
        return None  # header, separator, active-selection echo, or import slot

    # Cheap gate before reading all eight series: a real dataset has an
    # air-temperature series and coordinates. Empty user-defined template slots
    # carry the show-flag but none of these, so they are skipped here.
    latitude = _cell(row, _COL_LATITUDE)
    longitude = _cell(row, _COL_LONGITUDE)
    air_c = _read_series(row, 0, "air_c")
    if air_c is None or not isinstance(latitude, (int, float)) or not isinstance(longitude, (int, float)):
        return None

    series = {name: _read_series(row, index, name) for index, name in enumerate(_SERIES_ORDER)}

    key = str(_cell(row, _COL_KEY) or "").strip()
    name = str(_cell(row, _COL_NAME) or key).strip()

    return ClimateRecord(
        display_name=name,
        provider="phi",
        version=None,  # set by the seed routine via the dataset row
        station_id=key or None,
        phpp_codes=ClimatePhppCodes(
            country_code=str(_cell(row, _COL_COUNTRY) or "").strip(),
            region_code=str(_cell(row, _COL_REGION) or "").strip(),
            dataset_name=f"{key}-{name}" if key else name,  # the PHPP dropdown value
        ),
        location=ClimateLocation(
            latitude=float(latitude),
            longitude=float(longitude),
            site_elevation_m=_opt_float(_cell(row, _COL_ELEVATION)),
            hours_from_utc=round(float(longitude) / 15.0),
        ),
        climate=ClimateData(
            station_elevation_m=_opt_float(_cell(row, _COL_ELEVATION)) or 0.0,
            summer_daily_temperature_swing_k=_opt_float(_cell(row, _COL_SUMMER_SWING)) or 8.0,
            monthly_temps=ClimateMonthlyTemps(
                air_c=air_c,
                dewpoint_c=series["dewpoint_c"] or _ZEROS,
                sky_c=series["sky_c"] or _ZEROS,
                ground_c=_ZEROS,  # the PHPP library carries no monthly ground temps
            ),
            monthly_radiation=ClimateMonthlyRadiation(
                north=_require_series(series, "north", key),
                east=series["east"] or _ZEROS,
                south=series["south"] or _ZEROS,
                west=series["west"] or _ZEROS,
                glob=_require_series(series, "glob", key),
            ),
            peak_loads=_read_peak_loads(row),
        ),
        aux=ClimateAux(),  # the PHPP library has no Phius-style degree-hour / wind aux block
    )


def _read_series(row: tuple[object, ...], index: int, label: str) -> list[float] | None:
    """Read the ``index``-th 12-month block; ``None`` when wholly absent.

    A block must be all-numeric (used) or all-empty (``None``); a partially
    filled block is malformed and raises :class:`PhiParseError`.
    """
    start = _SERIES_START + index * _MONTHS
    cells = [_cell(row, start + offset) for offset in range(_MONTHS)]
    numbers = [cell for cell in cells if isinstance(cell, (int, float))]
    if len(numbers) == _MONTHS:
        return [float(value) for value in numbers]
    if not numbers:
        return None
    raise PhiParseError(f"series {label!r} has {len(numbers)}/{_MONTHS} numeric months (partial block).")


def _require_series(series: dict[str, list[float] | None], name: str, key: str) -> list[float]:
    """A series that every real dataset must carry (air/north/global)."""
    values = series[name]
    if values is None:
        raise PhiParseError(f"dataset {key!r} is missing the required {name!r} series.")
    return values


def _read_peak_loads(row: tuple[object, ...]) -> ClimatePeakLoads:
    """Read the four peak/design conditions; blank conditions default to zeros."""
    return ClimatePeakLoads(
        **{field: _read_peak(row, start, has_dewpoint) for field, start, has_dewpoint in _PEAK_BLOCKS}
    )


def _read_peak(row: tuple[object, ...], start: int, has_dewpoint: bool) -> ClimatePeakLoad:
    """One design condition: air temp + five radiation columns (+ dewpoint for cooling)."""
    return ClimatePeakLoad(
        temp_c=_opt_float(_cell(row, start)) or 0.0,
        rad_north=_opt_float(_cell(row, start + 1)) or 0.0,
        rad_east=_opt_float(_cell(row, start + 2)) or 0.0,
        rad_south=_opt_float(_cell(row, start + 3)) or 0.0,
        rad_west=_opt_float(_cell(row, start + 4)) or 0.0,
        rad_global=_opt_float(_cell(row, start + 5)) or 0.0,
        dewpoint_c=_opt_float(_cell(row, start + 6)) if has_dewpoint else None,
    )


def _cell(row: tuple[object, ...], col: int) -> object:
    """The 1-based ``col`` of a ``values_only`` row, or ``None`` past its end."""
    index = col - 1
    return row[index] if index < len(row) else None


def _opt_float(value: object) -> float | None:
    """A numeric cell as ``float``; ``None`` for blank / non-numeric cells."""
    return float(value) if isinstance(value, (int, float)) else None

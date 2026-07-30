"""CSV/XLSX serialization and saved-version export route tests."""

from __future__ import annotations

import csv
import io
from datetime import UTC, datetime
from pathlib import Path
from typing import Literal, cast
from uuid import UUID

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from features.access.capabilities import PROJECT_VIEW
from features.access.principals import UserPrincipal
from features.aperture_u_value.report import build_aperture_u_value_report
from features.aperture_u_value.report_csv import render_aperture_u_value_csv
from features.aperture_u_value.report_xlsx import render_aperture_u_value_xlsx
from features.aperture_u_value.units import (
    M2_TO_FT2,
    M_TO_FT,
    MM_TO_IN,
    W_M2K_TO_BTU_HRFT2F,
)
from features.auth.service import create_or_update_user
from features.project_document.document import ApertureTypeEntry
from features.projects.access import (
    ProjectAccess,
    require_project_view_access,
)
from features.projects.models import ProjectSummary
from main import app
from tests.test_aperture_u_value_parity import _element, _frames, _tables
from tests.test_apertures_mcp import _seed_aperture
from tests.test_mcp import clean_mcp_tables, create_project, signed_in_client

__all__ = ["clean_mcp_tables"]

GOLDEN_DIR = Path(__file__).with_name("goldens")
GENERATED_AT = datetime(2026, 7, 29, 18, 30, tzinfo=UTC)


def _export_report():
    tables = _tables()
    entry = ApertureTypeEntry(
        id="apt_export",
        name="Synthetic Export",
        row_heights_mm=[1000.0],
        column_widths_mm=[1200.0, 1000.0],
        elements=[
            _element("aptel_complete", column=0),
            _element(
                "aptel_unfinished",
                column=1,
                frames=_frames(top="pfrm_incomplete"),
            ),
        ],
    )
    tables.apertures = [entry]
    return build_aperture_u_value_report(
        project_id=UUID("00000000-0000-0000-0000-000000000001"),
        version_id=UUID("00000000-0000-0000-0000-000000000002"),
        source="version",
        project_name="Synthetic House",
        bt_number="TEST-01",
        version_label="Saved v1",
        tables=tables,
    )


@pytest.mark.parametrize(
    ("units", "golden_name"),
    [
        ("SI", "aperture-u-value-report-si.csv"),
        ("IP", "aperture-u-value-report-ip.csv"),
    ],
)
def test_csv_matches_synthetic_golden_with_bom_and_crlf(
    units: Literal["SI", "IP"],
    golden_name: str,
) -> None:
    rendered = render_aperture_u_value_csv(
        _export_report(),
        units=units,
    ).encode("utf-8")

    normalized = rendered.decode("utf-8-sig").replace("\r\n", "\n")
    assert normalized == (GOLDEN_DIR / golden_name).read_text(encoding="utf-8")
    assert rendered.startswith(b"\xef\xbb\xbf")
    assert b"\r\n" in rendered
    assert b"\n" not in rendered.replace(b"\r\n", b"")


def test_xlsx_pins_inputs_formulas_unfinished_rows_and_summary_refs() -> None:
    data = render_aperture_u_value_xlsx(
        _export_report(),
        units="SI",
        generated_at=GENERATED_AT,
    )
    workbook = load_workbook(io.BytesIO(data), data_only=False)
    detail = workbook["Window Units"]
    summary = workbook["Summary"]
    columns = {cell.value: cell.column for cell in detail[9] if isinstance(cell.value, str)}

    def ref(header: str, row: int = 10) -> str:
        return f"{get_column_letter(columns[header])}{row}"

    assert detail["B5"].value == GENERATED_AT.isoformat()
    assert "shared mullions are charged fully" in detail["B6"].value
    assert detail["B7"].value == ("Aperture U-w includes unfinished elements as U = 0.")
    assert detail[ref("Width [m]")].value == 1.2
    assert detail[ref("Top Frame Width [mm]")].value == 80.0
    assert detail[ref("Glazing U [W/(m²·K)]")].value == 0.7
    assert detail[ref("Top Ψ-g [W/(m·K)]")].value == 0.04
    assert "excluded from U-w" in next(header for header in columns if header.startswith("Top Ψ-install"))

    top_interior = ref("Top Interior Length [m]")
    left_width = ref("Left Frame Width [mm]")
    right_width = ref("Right Frame Width [mm]")
    assert detail[top_interior].value == (f"=G10-{left_width}/1000-{right_width}/1000")
    assert detail[ref("Top Center-Strip Area [m²]")].value == (f"=({ref('Top Frame Width [mm]')}/1000)*{top_interior}")
    assert detail[ref("Top Half-Corner A [m²]")].value == (
        f"=({ref('Top Frame Width [mm]')}/1000)*({left_width}/1000)/2"
    )
    for side in ("Top", "Right", "Bottom", "Left"):
        assert str(detail[ref(f"{side} Frame Area [m²]")].value).startswith("=SUM(")
        assert str(detail[ref(f"{side} Q-frame [W/K]")].value).startswith("=")
        assert str(detail[ref(f"{side} Q-spacer [W/K]")].value).startswith("=")

    element_u = str(detail[ref("Element U [W/(m²·K)]")].value)
    aperture_u = str(detail[ref("Aperture U-w [W/(m²·K)]")].value)
    aperture_shgc = str(detail[ref("Aperture SHGC [glazing-area weighted]")].value)
    assert element_u.startswith("=ROUND(")
    assert "SUMPRODUCT(" in aperture_u
    assert ",$BT$10:$BT$11)" in aperture_u
    assert "IFERROR" not in aperture_u
    assert "SUM($I$10:$I$11)" in aperture_u
    assert "IFERROR" not in aperture_shgc
    assert 'IF(SUMPRODUCT($BP$10:$BP$11,--ISNUMBER($M$10:$M$11))=0,""' in aperture_shgc
    assert "SUMPRODUCT($BP$10:$BP$11,$M$10:$M$11)" in aperture_shgc

    unfinished_row = 11
    assert detail[ref("Status", unfinished_row)].value == "UNFINISHED"
    assert detail[ref("Width [m]", unfinished_row)].value == 1.0
    assert detail[ref("Element Area [m²]", unfinished_row)].value == 1.0
    assert detail[ref("Glazing U [W/(m²·K)]", unfinished_row)].value == 0.7
    assert detail[ref("Top Frame Width [mm]", unfinished_row)].value == 80.0
    assert detail[ref("Top Ψ-g [W/(m·K)]", unfinished_row)].value is None
    assert detail[ref("Top Interior Length [m]", unfinished_row)].value == "UNFINISHED"
    assert detail[ref("Element U [W/(m²·K)]", unfinished_row)].value == "UNFINISHED"
    assert detail[ref("Aperture U-w [W/(m²·K)]", unfinished_row)].value == f"={ref('Aperture U-w [W/(m²·K)]')}"

    assert summary["C2"].value == (f"='Window Units'!{ref('Aperture Area [m²]')}")
    assert summary["D2"].value == (f"='Window Units'!{ref('Aperture U-w [W/(m²·K)]')}")
    assert summary["E2"].value == (f"='Window Units'!{ref('Aperture SHGC [glazing-area weighted]')}")


def test_xlsx_ip_inputs_and_rounding_formula_use_canonical_conversions() -> None:
    workbook = load_workbook(
        io.BytesIO(
            render_aperture_u_value_xlsx(
                _export_report(),
                units="IP",
                generated_at=GENERATED_AT,
            )
        ),
        data_only=False,
    )
    detail = workbook["Window Units"]
    columns = {cell.value: cell.column for cell in detail[9]}

    def value(header: str):
        return detail.cell(row=10, column=columns[header]).value

    assert value("Width [ft]") == pytest.approx(1.2 * M_TO_FT)
    assert value("Element Area [ft²]") == pytest.approx(1.2 * M2_TO_FT2)
    assert value("Top Frame Width [in]") == pytest.approx(80.0 * MM_TO_IN)
    assert value("Glazing U [Btu/(h·ft²·°F)]") == pytest.approx(0.7 * W_M2K_TO_BTU_HRFT2F)
    element_u_formula = value("Element U [Btu/(h·ft²·°F)]")
    aperture_u_formula = value("Aperture U-w [Btu/(h·ft²·°F)]")
    assert f"/{W_M2K_TO_BTU_HRFT2F},4)" in element_u_formula
    assert f"/{W_M2K_TO_BTU_HRFT2F},4)" in aperture_u_formula


def test_xlsx_preserves_available_inputs_on_non_positive_glazing_geometry() -> None:
    tables = _tables()
    entry = ApertureTypeEntry(
        id="apt_non_positive_export",
        name="Non-positive",
        row_heights_mm=[1000.0],
        column_widths_mm=[1000.0],
        elements=[
            _element(
                "aptel_non_positive_export",
                frames=_frames(
                    top="pfrm_wide",
                    right="pfrm_wide",
                    bottom="pfrm_wide",
                    left="pfrm_wide",
                ),
            )
        ],
    )
    tables.apertures = [entry]
    report = build_aperture_u_value_report(
        project_id=UUID("00000000-0000-0000-0000-000000000001"),
        version_id=UUID("00000000-0000-0000-0000-000000000002"),
        source="version",
        project_name="Synthetic",
        bt_number="TEST",
        version_label="Saved",
        tables=tables,
    )
    detail = load_workbook(
        io.BytesIO(
            render_aperture_u_value_xlsx(
                report,
                units="SI",
                generated_at=GENERATED_AT,
            )
        ),
        data_only=False,
    )["Window Units"]
    columns = {cell.value: cell.column for cell in detail[9]}

    assert detail.cell(10, columns["Top Frame Width [mm]"]).value == 600.0
    assert detail.cell(10, columns["Top Frame U [W/(m²·K)]"]).value == 1.0
    assert detail.cell(10, columns["Top Ψ-g [W/(m·K)]"]).value == 0.04
    assert detail.cell(10, columns["Top Frame Area [m²]"]).value == "UNFINISHED"
    assert detail.cell(10, columns["Element U [W/(m²·K)]"]).value == "UNFINISHED"


def test_xlsx_rollups_group_two_apertures_once_and_reuse_first_row() -> None:
    tables = _tables()
    tables.apertures = [
        ApertureTypeEntry(
            id="apt_group_a",
            name="Group A",
            row_heights_mm=[1000.0],
            column_widths_mm=[1200.0, 1000.0],
            elements=[
                _element("aptel_group_a1", column=0),
                _element("aptel_group_a2", column=1, glazing_id="pglz_b"),
            ],
        ),
        ApertureTypeEntry(
            id="apt_group_b",
            name="Group B",
            row_heights_mm=[800.0],
            column_widths_mm=[700.0],
            elements=[_element("aptel_group_b1", glazing_id="pglz_b")],
        ),
    ]
    report = build_aperture_u_value_report(
        project_id=UUID("00000000-0000-0000-0000-000000000001"),
        version_id=UUID("00000000-0000-0000-0000-000000000002"),
        source="version",
        project_name="Synthetic",
        bt_number="TEST",
        version_label="Saved",
        tables=tables,
    )
    workbook = load_workbook(
        io.BytesIO(
            render_aperture_u_value_xlsx(
                report,
                units="SI",
                generated_at=GENERATED_AT,
            )
        ),
        data_only=False,
    )
    detail = workbook["Window Units"]
    summary = workbook["Summary"]
    columns = {cell.value: cell.column for cell in detail[9]}
    u_column = get_column_letter(columns["Aperture U-w [W/(m²·K)]"])
    area_column = get_column_letter(columns["Aperture Area [m²]"])
    shgc_column = get_column_letter(columns["Aperture SHGC [glazing-area weighted]"])

    first_group_formula = str(detail[f"{u_column}10"].value)
    assert "$I$10:$I$11" in first_group_formula
    assert "$BT$10:$BT$11" in first_group_formula
    assert "$I$12" not in first_group_formula
    assert detail[f"{area_column}10"].value == "=SUM($I$10:$I$11)"
    assert detail[f"{area_column}11"].value == f"={area_column}10"
    assert detail[f"{u_column}11"].value == f"={u_column}10"
    assert detail[f"{shgc_column}11"].value == f"={shgc_column}10"
    second_group_formula = str(detail[f"{u_column}12"].value)
    assert "$I$12:$I$12" in second_group_formula
    assert "$BT$12:$BT$12" in second_group_formula
    assert summary["D2"].value == f"='Window Units'!{u_column}10"
    assert summary["D3"].value == f"='Window Units'!{u_column}12"


def test_csv_and_xlsx_neutralize_user_authored_formula_prefixes() -> None:
    tables = _tables()
    tables.project_frames[0].name = "=frame"
    tables.project_glazings[0].name = "@glazing"
    entry = ApertureTypeEntry(
        id="apt_formula_text",
        name="+aperture",
        row_heights_mm=[1000.0],
        column_widths_mm=[1000.0],
        elements=[_element("aptel_formula_text")],
    )
    entry.elements[0].name = "-element"
    tables.apertures = [entry]
    report = build_aperture_u_value_report(
        project_id=UUID("00000000-0000-0000-0000-000000000001"),
        version_id=UUID("00000000-0000-0000-0000-000000000002"),
        source="version",
        project_name="=project",
        bt_number="TEST",
        version_label="Saved",
        tables=tables,
    )

    csv_rows = list(csv.reader(io.StringIO(render_aperture_u_value_csv(report, units="SI").removeprefix("\ufeff"))))
    header = {name: index for index, name in enumerate(csv_rows[0])}
    assert csv_rows[1][header["project_name"]] == "'=project"
    assert csv_rows[1][header["aperture_name"]] == "'+aperture"
    assert csv_rows[1][header["element_name"]] == "'-element"
    assert csv_rows[1][header["glazing_name"]] == "'@glazing"
    assert csv_rows[1][header["top_frame_name"]] == "'=frame"

    detail = load_workbook(
        io.BytesIO(
            render_aperture_u_value_xlsx(
                report,
                units="SI",
                generated_at=GENERATED_AT,
            )
        ),
        data_only=False,
    )["Window Units"]
    columns = {cell.value: cell.column for cell in detail[9]}
    assert detail["B2"].value == "'=project"
    for header_name, expected in (
        ("Aperture Name", "'+aperture"),
        ("Element Name", "'-element"),
        ("Glazing Name", "'@glazing"),
        ("Top Frame", "'=frame"),
    ):
        cell = detail.cell(10, columns[header_name])
        assert cell.value == expected
        assert cell.data_type == "s"


def test_export_route_is_gated_defaults_ip_and_validates_query_values(
    clean_mcp_tables: None,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    editor = signed_in_client()
    project = create_project(editor)
    project_id = cast(str, project["id"])
    version_id = cast(str, project["active_version_id"])
    _seed_aperture(version_id)
    url = f"/api/v1/projects/{project_id}/versions/{version_id}/apertures/u-values/report/export"

    csv_response = editor.get(
        url,
        params={"format": "csv"},
        headers={"Origin": "http://localhost:5173"},
    )
    assert csv_response.status_code == 200
    assert csv_response.headers["content-type"].startswith("text/csv")
    assert "aperture-u-values-IP-" in csv_response.headers["content-disposition"]
    assert "Content-Disposition" in csv_response.headers["access-control-expose-headers"]
    assert csv_response.content.startswith(b"\xef\xbb\xbf")

    xlsx_response = editor.get(
        url,
        params={"format": "xlsx", "units": "SI"},
    )
    assert xlsx_response.status_code == 200
    assert xlsx_response.content.startswith(b"PK")
    assert xlsx_response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert editor.get(url, params={"format": "pdf"}).status_code == 422
    assert (
        editor.get(
            url,
            params={"format": "csv", "units": "metric"},
        ).status_code
        == 422
    )
    assert TestClient(app).get(url, params={"format": "csv"}).status_code == 401

    limited_user = create_or_update_user(
        email="limited@example.com",
        display_name="Limited User",
        password="password",
    )
    project_summary = ProjectSummary.model_validate({field: project[field] for field in ProjectSummary.model_fields})
    limited_access = ProjectAccess(
        project_id=project_summary.id,
        mode="view",
        principal=UserPrincipal(user=limited_user),
        project=project_summary,
    )
    monkeypatch.setattr(
        "features.projects.access.capabilities_for",
        lambda _principal: frozenset({PROJECT_VIEW}),
    )
    app.dependency_overrides[require_project_view_access] = lambda: limited_access
    try:
        assert TestClient(app).get(url, params={"format": "csv"}).status_code == 403
    finally:
        app.dependency_overrides.pop(require_project_view_access, None)

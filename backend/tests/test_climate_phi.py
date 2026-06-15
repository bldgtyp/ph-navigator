"""Tests for the PHI/PHPP workbook importer (`features.climate.importers.phi`).

The real PHI/PHPP workbook is licensed and never committed (D-CS-6), so these
build a **synthetic** ``Climate`` worksheet in ``tmp_path`` with fabricated
numbers placed at the documented column positions. The fixture doubles as an
executable spec of the recovered column map: if the importer's column
constants drift, the golden assertions here fail.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from features.climate.importers import get_provider, provider_names, resolve_version
from features.climate.importers.phi import PhiParseError, iter_phi_records, parse_phi_workbook
from features.climate.record import ClimateRecord

# Each 12-month series, first column letter, per the recovered layout.
_SERIES_COLS = {
    "air": "W",
    "north": "AI",
    "east": "AU",
    "south": "BG",
    "west": "BS",
    "glob": "CE",
    "dewpoint": "CQ",
    "sky": "DC",
}
# Peak/design conditions: field → (first column, dewpoint column or None).
_PEAK_COLS = {
    "heat_load_1": ("DO", None),
    "heat_load_2": ("DU", None),
    "cooling_load_1": ("EA", "EG"),
    "cooling_load_2": ("EH", "EN"),
}


def _put(ws: Worksheet, col: str, row: int, value: object) -> None:
    ws[f"{col}{row}"] = value


def _put_block(ws: Worksheet, col: str, row: int, values: list[float]) -> None:
    start = column_index_from_string(col)
    for offset, value in enumerate(values):
        ws.cell(row=row, column=start + offset, value=value)


def _put_peak(ws: Worksheet, field: str, row: int, temp: float, rad: list[float], dewpoint: float | None) -> None:
    start_col, dew_col = _PEAK_COLS[field]
    _put_block(ws, start_col, row, [temp, *rad])
    if dew_col is not None and dewpoint is not None:
        _put(ws, dew_col, row, dewpoint)


def _ramp(base: float) -> list[float]:
    """A fabricated 12-month series, ``base`` … ``base + 11``."""
    return [base + month for month in range(12)]


def _write_station(
    ws: Worksheet, row: int, *, show: str, key: str, name: str, with_sky: bool, with_peaks: bool
) -> None:
    _put(ws, "K", row, show)
    _put(ws, "I", row, key)
    _put(ws, "N", row, "ZZ")
    _put(ws, "O", row, "Testland")
    _put(ws, "P", row, name)
    _put(ws, "S", row, 12.5)
    _put(ws, "T", row, -45.0)
    _put(ws, "U", row, 100.0)
    _put(ws, "V", row, 9.0)
    _put_block(ws, _SERIES_COLS["air"], row, _ramp(1.0))
    _put_block(ws, _SERIES_COLS["north"], row, _ramp(10.0))
    _put_block(ws, _SERIES_COLS["east"], row, _ramp(20.0))
    _put_block(ws, _SERIES_COLS["south"], row, _ramp(30.0))
    _put_block(ws, _SERIES_COLS["west"], row, _ramp(40.0))
    _put_block(ws, _SERIES_COLS["glob"], row, _ramp(50.0))
    _put_block(ws, _SERIES_COLS["dewpoint"], row, _ramp(-5.0))
    if with_sky:
        _put_block(ws, _SERIES_COLS["sky"], row, _ramp(-15.0))
    if with_peaks:
        _put_peak(ws, "heat_load_1", row, -12.0, [1, 2, 3, 4, 5], None)
        _put_peak(ws, "cooling_load_1", row, 30.0, [11, 12, 13, 14, 15], 18.0)


def _synthetic_workbook(path: Path) -> Path:
    """A workbook with one full dataset, one sparse dataset, an echo, a blank slot."""
    workbook = openpyxl.Workbook()
    ws = workbook.active
    assert ws is not None
    ws.title = "Climate"
    # Excluded: the active-selection echo carries data but show-flag "-".
    _write_station(ws, 10, show="-", key="ZZ9999", name="Echo", with_sky=True, with_peaks=True)
    # Station A: full series + peak loads + dewpoint/sky.
    _write_station(ws, 11, show="x", key="ZZ0001a", name="Synthville", with_sky=True, with_peaks=True)
    # Station B: no sky series, no peak loads (both default).
    _write_station(ws, 12, show="x", key="ZZ0002b", name="Dryville", with_sky=False, with_peaks=False)
    # Skipped: an empty user-defined template slot (show-flag set, no data).
    _put(ws, "K", 13, "x")
    _put(ws, "I", 13, "ud---01")
    workbook.save(path)
    return path


@pytest.fixture()
def workbook(tmp_path: Path) -> Path:
    return _synthetic_workbook(tmp_path / "phi_synthetic.xlsx")


def _by_id(path: Path) -> dict[str | None, ClimateRecord]:
    return {record.station_id: record for record in parse_phi_workbook(path)}


# --- Selection: which rows become datasets ----------------------------------


def test_parses_only_real_datasets(workbook: Path) -> None:
    records = list(parse_phi_workbook(workbook))
    # The echo row (show="-") and the empty template slot are both excluded.
    assert sorted(r.station_id for r in records) == ["ZZ0001a", "ZZ0002b"]


# --- Golden decode of the full column map -----------------------------------


def test_golden_decode_full_station(workbook: Path) -> None:
    record = _by_id(workbook)["ZZ0001a"]

    assert record.provider == "phi"
    assert record.version is None  # assigned by the seed routine
    assert record.display_name == "Synthville"
    assert record.phpp_codes.country_code == "ZZ"
    assert record.phpp_codes.region_code == "Testland"
    assert record.phpp_codes.dataset_name == "ZZ0001a-Synthville"

    assert record.location.latitude == 12.5
    assert record.location.longitude == -45.0
    assert record.location.site_elevation_m == 100.0
    assert record.location.hours_from_utc == -3  # round(-45 / 15)
    assert record.climate.summer_daily_temperature_swing_k == 9.0

    assert record.climate.monthly_temps.air_c == _ramp(1.0)
    assert record.climate.monthly_temps.dewpoint_c == _ramp(-5.0)
    assert record.climate.monthly_temps.sky_c == _ramp(-15.0)
    assert record.climate.monthly_radiation.north == _ramp(10.0)
    assert record.climate.monthly_radiation.east == _ramp(20.0)
    assert record.climate.monthly_radiation.south == _ramp(30.0)
    assert record.climate.monthly_radiation.west == _ramp(40.0)
    assert record.climate.monthly_radiation.glob == _ramp(50.0)

    heat_1 = record.climate.peak_loads.heat_load_1
    assert (heat_1.temp_c, heat_1.rad_north, heat_1.rad_global) == (-12.0, 1.0, 5.0)
    cooling_1 = record.climate.peak_loads.cooling_load_1
    assert (cooling_1.temp_c, cooling_1.rad_north, cooling_1.rad_global) == (30.0, 11.0, 15.0)
    assert cooling_1.dewpoint_c == 18.0  # cooling conditions carry a trailing dewpoint


def test_absent_optional_series_and_peaks_default(workbook: Path) -> None:
    record = _by_id(workbook)["ZZ0002b"]

    # Sky was omitted entirely → a zero vector (mirrors the Phius importer).
    assert record.climate.monthly_temps.sky_c == [0.0] * 12
    # Dewpoint was present, so it is preserved.
    assert record.climate.monthly_temps.dewpoint_c == _ramp(-5.0)
    # No design columns → every peak load defaults to zeros.
    assert record.climate.peak_loads.heat_load_1.temp_c == 0.0
    assert record.climate.peak_loads.cooling_load_1.dewpoint_c is None


# --- Error handling ---------------------------------------------------------


def test_partial_series_raises(tmp_path: Path) -> None:
    path = tmp_path / "partial.xlsx"
    workbook = openpyxl.Workbook()
    ws = workbook.active
    assert ws is not None
    ws.title = "Climate"
    _write_station(ws, 11, show="x", key="ZZ0003c", name="Brokeville", with_sky=True, with_peaks=False)
    ws["AB11"] = None  # blow a hole in the middle of the air-temperature series
    workbook.save(path)

    with pytest.raises(PhiParseError, match="partial block"):
        list(parse_phi_workbook(path))


def test_missing_climate_sheet_raises(tmp_path: Path) -> None:
    path = tmp_path / "wrong.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "NotClimate"
    workbook.save(path)

    with pytest.raises(PhiParseError, match="no 'Climate' worksheet"):
        list(parse_phi_workbook(path))


# --- Tree walk + registry wiring --------------------------------------------


def test_iter_phi_records_walks_tree(tmp_path: Path) -> None:
    nested = tmp_path / "phi" / "10.6"
    nested.mkdir(parents=True)
    _synthetic_workbook(nested / "library.xlsx")

    records = list(iter_phi_records(tmp_path))
    assert sorted(r.station_id for r in records) == ["ZZ0001a", "ZZ0002b"]


def test_phi_provider_is_registered() -> None:
    assert "phi" in provider_names()
    assert resolve_version("phi", None) == "10.6"
    assert get_provider("phi").label_for("10.6") == "PHI 10.6"
